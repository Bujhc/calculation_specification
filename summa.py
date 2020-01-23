import glob
import os
import pathlib

import openpyxl

from pathlib import Path

from openpyxl import Workbook, load_workbook, cell



summa = []

os.chdir(r"D:\Temp\Сколково_Бизнес парк\Сколково Бизнес парк\CTCS    (Светофоры)")

for filename in Path().rglob('*Спецификация*.xlsx'):

    filepath = os.path.abspath(filename)
    print(filepath)
    workbook = load_workbook(filename=str(filepath), data_only=True)

    sheet = workbook.active
    
    UPDATE = {'ИТОГО':'ИТОГО'}

    for i in range(2, sheet.max_row):
        articleName = sheet.cell(row=i, column=2).value
        if articleName in UPDATE:
            summa.append(sheet.cell(row=i, column=9).value)
            print(summa)

print(sum(summa))            
           

os.chdir(r"C:\Users\bujhc\Project\excel")   

excel_file = Workbook()
excel_sheet = excel_file.create_sheet(title='Result', index=0)     
next_row =1
for items in summa:
    
    excel_sheet.cell(column=2, row=next_row, value=items)
        
    next_row +=1
    

excel_file.save('rezult_sum.xlsx')
