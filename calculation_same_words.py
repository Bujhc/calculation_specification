import glob
import os
import pathlib

import openpyxl

from pathlib import Path
from openpyxl import Workbook, load_workbook, cell


rezult = {}
list_of_signals = []

os.chdir(r"D:\Temp\Сколково_Бизнес парк\Сколково Бизнес парк\LC     (управление освещением)")
i = 0
for filename in Path().rglob('Список* точек*.xlsx'):
    
    filepath = os.path.abspath(filename)
    

    wb = load_workbook(filename=str(filepath))

    ws_list = len(wb.sheetnames)

    for number_list_wb in range(0,ws_list):


        sheet = wb.worksheets[number_list_wb]

        
        

        list_of_dictionary = ["Name", "Analog Input", "Binary Input", "Analog Output", "Binary Output"]

        Analog_Input=0
        Binary_Input=0
        Analog_Output=0
        Binary_Output=0

        for x in sheet.iter_rows(min_row=1, min_col=1, max_row=150, max_col=6, values_only=True):

            Analog_Input += (x.count("Analog Input"))
            Binary_Input += (x.count("Binary Input"))
            Analog_Output += (x.count("Analog Output"))
            Binary_Output += (x.count("Binary Output"))
        i +=1
        list_of_signals = [filename, Analog_Input,Binary_Input,Analog_Output,Binary_Output]
        rezult[i] = dict(zip(list_of_dictionary,list_of_signals))
    
print(ws_list)

os.chdir(r"C:\Users\bujhc\Project\excel")   

excel_file = Workbook()
excel_sheet = excel_file.create_sheet(title='Point_of_controller', index=0)

header = ("Name", "Analog Input", "Binary Input", "Analog Output", "Binary Output")
excel_sheet.append(header)

next_row=2

for items in rezult.values():
    
    excel_sheet.cell(column=2, row=next_row, value=items["Analog Input"])
    excel_sheet.cell(column=3, row=next_row, value=items["Binary Input"])
    excel_sheet.cell(column=4, row=next_row, value=items["Analog Output"])
    excel_sheet.cell(column=5, row=next_row, value=items["Binary Output"])
    
    next_row +=1
    

excel_file.save('rezult.xlsx')