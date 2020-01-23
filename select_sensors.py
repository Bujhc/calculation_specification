import glob
import os
import pathlib

import openpyxl

from pathlib import Path

from openpyxl import Workbook, load_workbook, cell





os.chdir(r"D:\Temp\Сколково_Бизнес парк\Сколково Бизнес парк\FCU&Heaters    (фанкойлы и нагреватели )")

for filename in Path().rglob('*.xlsx'):

    filepath = os.path.abspath(filename)
    print(filepath)
    workbook = load_workbook(filename=str(filepath))

    sheet = workbook.active


    UPDATE = {'Датчик перепада давления на фильтре': 'P233A-4-PKC',
            'Контроллер UNI-B-X': 'TUC0311-2',
            'Комнатный контроллер/ Space Smartstat': 'TUC0311-2',
            'Датчик температуры комнатный': 'TM-2140-0000',
            'Датчик перепада давления на вентиляторе': 'P233A-10-PKC',
            'Датчик температуры воды': 'TS-6330S-000',
            'Датчик температуры воздуха': 'TS-6330D-B10',
            'Датчик температуры воздуха внутри помещения': 'TM-1140-0000',
            'Датчик защиты от замораживания': '270XT-95008',
            'Привод заслонок (с пруж.возвратом)': 'M9208-BGC-1',
            'Привод заслонок (с пруж. возвратом)': 'M9208-BGC-1',
            'Комплект монтажный': '9999',
            'Комплект маркировочный': '8888'
            }


    for i in range(2, sheet.max_row):
        articleName = sheet.cell(row=i, column=2).value
        if articleName in UPDATE:
            sheet.cell(row=i, column=3).value = UPDATE[articleName]

    workbook.save(filepath)

