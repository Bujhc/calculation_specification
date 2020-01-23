import glob
import os
import pathlib

import openpyxl

from pathlib import Path

from openpyxl import Workbook, load_workbook, cell





os.chdir(r"D:\Temp\Сколково_Бизнес парк\Сколково Бизнес парк\CTCS    (Светофоры)")

for filename in Path().rglob('Спецификация*.xlsx'):

    filepath = os.path.abspath(filename)
    print(filepath)
    workbook = load_workbook(filename=str(filepath))

    workbook.guess_types = True

    sheet = workbook.active
    


    UPDATE = {'P233A-4-PKC': 4942,
            'TM-2140-0000': 4700,
            'TUC0311-2': 18500,
            'P233A-10-PKC': 4942,
            'TS-6330S-000': 5282,
            'TS-6330D-B10': 3680,
            'TM-1140-0000': 4499,
            '270XT-95008': 4516,
            'M9208-BGC-1': 21016,
            'TWA-Z (NC)': 3300,
            'AME 435 QM': 33000,
            '8727.0201': 2500,

            'AME-25SD': 59000,
            'AME-10': 31950,
            'САУ-М6': 3360,
            'R10': 3388,
            'M32': 26371,
            'PB2': 2896,
            'ИО-102-6(СМК-6)': 395,
            'СКАТ 1200И7': 6356,
            'LL-C-V': 9700,
            'SPHINX 104-360/2 AP': 9500,

            'RIGI CO0 L42M': 12351,
            62920: 68,
            91920: 63,
            51120: 6,
            54920: 13,
            50420: 13,
            53700: 101,
            'PLUG-8P8C-UV-C6': 10,
            'UTP4-C6a-SOLID-LSZH-GY': 101,
            'UTP4-C6-PATCH-NCR-GY': 101,
            'STP4-C6-SOLID-INDOOR-FR-PVC': 125,
            'STP4-C6-PATCH-INDOOR': 125,
            'KLM 2 x 0.8': 89,
            'KLM 4 x 0.8': 102,
            'KLM 2x0,8': 89,
            'KLM 4x0,8': 102,
            'KLM 2x0.8': 89,
            'KLM 4x0.8': 102,
            'ШВВП 3х0,5': 76,
            'J-Y(St)Y 4x2x0,8': 76,
            'КДВВГ 4х0,25': 76,
            'КПСнг-FRLS 1x2x1': 76,
            'ВВГнг-LS 4x1,5': 112,
            'ВВГнг-LS 4x6': 266,
            'ВВГнг-LS 5x1,5': 121,
            'ВВГнг-LS 4x2,5': 121,
            'ВВГнг-LS 3x2,5': 88,
            'ВВГнг-LS 3x1,5': 88,
            'ВВГнг-LS 4x4': 140,
            'ВВГнг-LS 2x1.5': 88,
            'ВВГЭнг-LS 4x4': 200,
            '8888': 1450,
            '9999': 4100,
            'ВВГнг-LS 3x1,0': 112
            }


    for i in range(2, sheet.max_row):
        articleName = sheet.cell(row=i, column=3).value
        if articleName in UPDATE:
            sheet.cell(row=i, column=8).value = UPDATE[articleName]
            
            cell_multiply = '=PRODUCT(G{0}:H{0})'.format(i)
            sheet.cell(row=i, column=9).value = cell_multiply

    workbook.save(filepath)