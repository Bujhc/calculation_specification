import glob
import os
import pathlib


from pathlib import Path

from openpyxl import Workbook, load_workbook, cell

os.chdir(r"C:\Users\bujhc\Project\excel")   


workbook = load_workbook(filename='Спецификация.xlsx', data_only=True)

sheet = workbook.active

UPDATE = {'ИТОГО':'ИТОГО'}

summa = []


for i in range(2, sheet.max_row):
    articleName = sheet.cell(row=i, column=2).value
    if articleName in UPDATE:
        print(i)
        summa.append(sheet.cell(row=i, column=9).value)

print(summa)